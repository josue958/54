import openpyxl

wb = openpyxl.load_workbook('assets/template.xlsx')
ws = wb.active

# Delete all rows from row 4 to the max row
max_row = ws.max_row
if max_row >= 4:
    ws.delete_rows(4, max_row - 3)

# Clear the specific title in row 2 since it's just a generic template
ws.cell(row=2, column=1).value = "PLANTILLA EN BLANCO"

wb.save('assets/template.xlsx')
print("Template cleared successfully.")
